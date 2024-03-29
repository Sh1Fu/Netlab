import logging
from datetime import datetime
from json import JSONDecodeError, loads
from os import makedirs
from os.path import exists
from shutil import make_archive, move
from time import sleep, strftime, strptime
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import urlretrieve

from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from pytz import timezone
from requests import get
from requests.exceptions import RequestException
from tqdm import tqdm


class DownloadImage:
    def __init__(self, file_name: str, creds: Any) -> None:
        self.LOG_FILE = "./out/%s.log" % strftime("%Y%m%d-%H%M")
        self.LOCAL_TIMEZONE = timezone("Europe/Moscow")
        self.AUTH_URL = "http://services.netlab.ru/rest/authentication/token.json?"
        self.PROXY = None
        self.file_name = file_name
        self.msg = ""
        self.creds = creds
        (self.token, self.live_time) = self.auth_token()
        if not exists("./images/"):
            print("[+] Make images/ dir to Netlab images..")
            makedirs("./images/")
        logging.basicConfig(filename=self.LOG_FILE, level=logging.INFO,
                            format='%(asctime)s.%(msecs)03d %(levelname)s %(module)s - %(funcName)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        logging.info("[+] Check requests with images..\n")

    def auth_token(self) -> tuple:
        '''
        API function. Take token from Netlab API\n
        ``creds`` - actual json credential data\n
        Return value: tuple
        * ``token`` - API token
        * ``live_time`` - raw time of token live
        '''
        req = get(self.AUTH_URL, self.creds)
        data_json = loads(req.text[req.text.find("& {") + 2:])
        token, live_time = data_json["tokenResponse"]["data"]["token"], data_json["tokenResponse"]["data"]["expiredIn"]
        return (token, live_time)

    def token_is_alive(self, time: str) -> bool:
        '''
        Check token live time\n
        Response time example:\n
        ``time`` - ``06.04.2017 13:15``
        '''
        expired_in = strptime(time, "%d.%m.%Y %H:%M")
        current_time = strptime(strftime("%d.%m.%Y %H:%M"), "%d.%m.%Y %H:%M")
        return 1 if current_time <= expired_in else 0

    def check_time(self) -> bool:
        '''
        Checks if the current work time falls within the Netlab manager work schedule\n
        ``Work Time`` --> 09:00 - 18:00
        '''
        now = datetime.now(self.LOCAL_TIMEZONE)
        work_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        work_end = now.replace(hour=23, minute=59, second=59, microsecond=0)
        return True if (work_start <= datetime.now(self.LOCAL_TIMEZONE) <= work_end) else False

    def scrap_proxy(self) -> str or None:
        '''
        Take all free proxy servers from free proxy list
        Resources:\n
        * free-proxy-list.net
        * httpstatus.io

        Check if response to http proxy has status code 200 and append it to clean_proxy list
        '''
        proxy_list = list()
        response = get('https://free-proxy-list.net/')
        proxy_table = BeautifulSoup(response.text, 'html.parser').find('table')
        proxy_html_raw = proxy_table.find_all("tr")
        for proxy_row in proxy_html_raw:
            td_tag = proxy_row.find_all('td')
            if len(td_tag) == 8:
                proxy_list.append(td_tag[0].get_text() + ":" + td_tag[1].get_text()) if proxy_row.find_all(
                    class_="hx")[0].get_text() == "no" and len(proxy_list) < 100 else None
        for bad_proxy in proxy_list:
            try:
                get("http://services.netlab.ru/rest/authentication/token.json",
                    proxies={'http': f'http://{bad_proxy}'}, timeout=1)
                return bad_proxy
            except RequestException:
                logging.warning(f"{bad_proxy} is not connected to Netlab")
                continue
        return None

    def take_image(self, id: str) -> str:
        '''
        API function. Take json with image's urls from Netlab API.
        Return value: image url or blank string
        '''

        headers = {
            "User-Agent": "%s" % UserAgent().random,
            'accept': "*/*",
            'cache-control': "no-cache"
        }
        data = None
        if not self.token_is_alive(self.live_time):
            (self.token, self.live_time) = self.auth_token()
        if self.check_time():
            try:
                response = get("http://services.netlab.ru/rest/catalogsZip/goodsImages/%s.json?oauth_token=%s" %
                               (id, self.token), headers=headers, proxies={'http': self.PROXY})
            except HTTPError or RequestException:
                self.msg = "NetworkError: Problems with proxy % on XML_ID %s" % (
                    self.PROXY if self.PROXY else "Main Network", id)
                logging.error(msg=self.msg)
                sleep(5)
                response = get("http://services.netlab.ru/rest/catalogsZip/goodsImages/%s.json?oauth_token=%s" %
                               (id, self.token), headers=headers, proxies=None)
            try:
                data = loads(response.text[response.text.find("& {") + 2:])
                if data['entityListResponse']['data'] is not None:
                    return data['entityListResponse']['data']['items'][0]['properties']['Url']
            except JSONDecodeError:
                self.msg = f"JSONDecodeError: Problems with response {data} on the XML_ID {id}"
                logging.error(msg=self.msg)
        else:
            self.msg = "Working time is over, waiting for the next day to start"
            logging.info(msg=self.msg)
            while not self.check_time():
                sleep(60)
        return ""

    def xlsx_work(self, with_proxy: bool) -> None:
        '''
        Main active function. Edit xlsx file. Add image's name to first unused column. Download first product's image\n
        You can choose whether to work with or without a proxy.\n
        IMPORTANT: Working with a proxy can significantly increase code time, as using a proxy is an unnecessary comparison.\n
        Moreover, open and free proxy servers are used 
        '''
        self.PROXY = self.scrap_proxy() if with_proxy else None
        print(self.PROXY)
        wb = load_workbook(
            filename=f"./price_lists/{self.file_name}", read_only=False)
        active_sh = wb.active
        sheet_length = active_sh.max_row
        current_column = active_sh.max_column + 1
        active_sh.cell(row=1, column=current_column).value = "Картинка"
        for i in tqdm(range(2, sheet_length + 1, 1)):
            XML_ID = active_sh["A%d" % i].value
            product_info = self.take_image(XML_ID)
            if product_info != "":
                active_sh.cell(
                    row=i, column=current_column).value = str(XML_ID) + ".jpg"
                try:
                    urlretrieve(
                        product_info, filename="./images/%s.jpg" % XML_ID)
                    sleep(0.2)
                except URLError or HTTPError or RequestException:
                    wb.save("./price_lists/images.xlsx")
                    logging.exception(f"Network Error on the row {i}: ")
                    sleep(30)
                    continue
        wb.save("./price_lists/images.xlsx")

    def sort_files(self, delit: int, files: list) -> None:
        '''
        Split all images to different dirs. Make zip archives after that
        '''
        a = files
        for i in range(1, (len(files) // delit) + 1):
            dir_name = f'./images/images-{i}'
            if not exists(dir_name):
                makedirs(dir_name)
            for j in range(delit):
                move(f"./images/images/{a[j]}", f"{dir_name}")
                print(f"move {a[j]} into {dir_name}")
            make_archive(f"images-{i}.zip", 'zip', f'./images/images-{i}/')
            a = a[delit:]

    def images_zip(self) -> None:
        '''
        Create zip with images to NetLab
        '''
        make_archive("./price_lists/images", 'zip', "./images/")
        print("[+] Zip file is created..")
