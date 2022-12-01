import logging
from csv import writer
from json import dump, loads
from os import makedirs, remove
from os.path import exists
from time import sleep, strftime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle
from requests import get
from tqdm import tqdm

from src.Netlab_UpdatePrice import UpdatePrice


class TakePrice(UpdatePrice):
    def __init__(self, auth_url: str, file_name: str) -> None:
        self.LOG_FILE = "./out/%s.log" % strftime("%Y%m%d-%H%M")
        self.CSV_NAME = "price_update_tmp.csv"
        self.auth_url = auth_url
        self.file_name = file_name
        self.diction = dict()
        self.current_row = 1
        self.token = None
        super().__init__(file_name=self.file_name)
        makedirs("./out/") if not exists("./out/") else None
        logging.basicConfig(filename=self.LOG_FILE, level=logging.INFO,
                            format='%(asctime)s.%(msecs)03d %(levelname)s %(module)s - %(funcName)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        logging.info("[+] Starting price update process..")
        if not exists("./price_lists/"):
            makedirs("./price_lists/")
            logging.info("[+] Make out result dir with updated price files..")

    def auth_token(self, creds: Any) -> None:
        '''
        API function. Take token from Netlab API\n
        ``creds`` - actual json credential data\n
        Return value: tuple
        * ``token`` - API token
        * ``live_time`` - raw time of token live
        '''
        req = get(self.auth_url, creds)
        data_json = self.prepare_json(req.text)
        self.token= data_json["tokenResponse"]["data"]["token"]

    def prepare_json(self, data: str) -> Any:
        '''
        Transform response from Netlab to Python dict\n
        ``data`` - response from all Netlab API functions
        '''
        return loads(data[data.find("& {") + 2:])

    def catalog_names(self, token) -> Any:
        '''
        API function. Take catalog (json object) from Netlab API. (catalog.json file)
        '''
        data = get(
            "http://services.netlab.ru/rest/catalogsZip/Прайс-лист.json?oauth_token=%s" % token)
        data_json = self.prepare_json(data.text)
        return data_json

    def find_name(self, json_object: list, id: str) -> str:
        return [obj for obj in json_object if obj['id'] == id][0]['name']

    def take_price(self, PRICE_TYPE: int, mode: int) -> None:
        '''
        API function. Take all products from category.
        '''
        catalog_json = self.catalog_names(self.token)
        with open("catalog.json", "w+") as catalog_file:
            dump(catalog_json, catalog_file)
        self.diction = self.update_list(self.diction)
        wb = Workbook()
        active_sheet = wb.active
        self.init_default_xlsx(active_sheet=active_sheet) if PRICE_TYPE == 0 else self.init_main_xlsx(
            active_sheet=active_sheet)
        try:
            catalog_json["catalogResponse"]["data"]["category"].index("Услуги и Получи!Фонд")
        except ValueError:
            logging.info("'Услуги и Получи!Фонд' category not in catalog")
        for subcatalog in tqdm(catalog_json["catalogResponse"]["data"]["category"]):
            products = get(
                "http://services.netlab.ru/rest/catalogsZip/Прайс-лист/%s.json?oauth_token=%s" % (subcatalog["id"], self.token))
            products = self.prepare_json(products.text)
            try:
                self.product_take(
                    PRICE_TYPE, products['categoryResponse']['data']['goods'], active_sheet, subcatalog["id"], mode)
            except BaseException as e:
                logging.error(f'Error: {str(e)}')
            else:
                continue
            sleep(0.2)
        wb.save(f"./price_lists/{self.file_name}")
        print("[+] Default price without images in result dir!")
        remove('catalog.json')

    def csv_save(self, file_name: str) -> None:
        '''
        Save new xlsx file as csv file with delimiter = ";"\n
        Change csv file name in constant ``CSV_NAME``
        '''
        wb = load_workbook(filename=file_name, read_only=False)
        active_sh = wb.active
        style = NamedStyle(name="style")
        style.font = Font(name="Arial")
        wb.add_named_style(style)
        with open(f"./price_lists/{self.CSV_NAME}", 'w', newline="") as result_file:
            csv_writer = writer(result_file, delimiter=";", quotechar='"')
            for row in active_sh.iter_rows():
                csv_writer.writerow([cell.value for cell in row])
        print("[+] Price list in csv format is ready!")
